import warnings
import xlsxwriter
import time
from openpyxl import load_workbook

def get_results(data, predictions, categories):
	'''
	Gets micro and macro F1 scores, as well as whether there are no tp, 
	precision, recall and F1 for each class
	'''
	state = get_states(data, predictions, categories)
	micro_f1, macro_f1 = evaluate(state, categories)	
	d = dict()
	for c in categories:
		b = True
		if state[c][0] > 0:
			b = False
		d[c] = [precision(state[c][0], state[c][1]), recall(state[c][0], state[c][3]), f1(state[c][0], state[c][1], state[c][3]), b]
	return micro_f1, macro_f1, d
	
def evaluate(state, categories):
	macro_f1 = sum(f1(state[c][0], state[c][1], state[c][3]) for c in categories)/len(categories)
	tp = sum(state[c][0] for c in categories)
	fp = sum(state[c][1] for c in categories)
	fn = sum(state[c][3] for c in categories)
	micro_f1 = f1(tp, fp, fn)
	return micro_f1, macro_f1

def get_states(data, predictions, categories):
	'''
	Receives a list of sentence-class tuples with the actual classes and 
	a list with the prediction for each index, and calculates the number 
	of tp, fp, tn and fn for each class.
	param [([str], str)] data
	param [str] predictions
	returns dict(str, [int, int, int, int])
	'''
	if len(data) != len(predictions):
		raise TypeError('Lists should be just equally long')
	state = dict()
	for c in categories:
		state[c] = [0.0, 0.0, 0.0, 0.0]
	for i in range(len(data)):
		if predictions[i] == data[i][1]:
			# tp
			state[predictions[i]][0] += 1
		else:
			# fp for predicted class
			state[predictions[i]][1] += 1
			# fn for real class
			state[data[i][1]][3] += 1
	# Add true negatives
	for c in categories:
		state[c][2] = len(predictions)-sum(state[c])
	return state

def precision(tp, fp):
	if tp+fp > 0:
		return tp/(tp+fp)
	else:
		return 0

def recall(tp, fn):
	if tp+fn > 0:
		return tp/(tp+fn)
	else:
		return 0

def f1(tp, fp, fn):
	if (precision(tp, fp)+recall(tp, fn)) > 0:
		return (2*precision(tp, fp)*recall(tp, fn))/(precision(tp, fp)+recall(tp, fn))
	else:
		return 0

def print_results(micro_f1, macro_f1, d, categories, exp_name=''):
	'''
	Prints evaluation in an excel table in the 'results' directory
	'''
	file_name = 'results/'+exp_name+'.xlsx'#+'_'+time.strftime("%d-%H-%M")+'.xlsx'
	workbook = xlsxwriter.Workbook(file_name)
	worksheet = workbook.add_worksheet()
	worksheet.set_landscape()
	worksheet.set_margins(left=0.3, right=0.3, top=0.3, bottom=0.3)
	bold = workbook.add_format({'bold': True})
	row = 0
	col = 0
	worksheet.write(row, col, 'Micro F1', bold)
	worksheet.write(row+1, col, micro_f1)
	col += 1
	worksheet.write(row, col, 'Macro F1', bold)
	worksheet.write(row+1, col, macro_f1)
	row += 3
	worksheet.write(row+1, 0, 'Precision', bold)	
	worksheet.write(row+2, 0, 'Recall', bold)	
	worksheet.write(row+3, 0, 'F1', bold)	
	col = 1
	f = workbook.add_format()
	f.set_font_color('#006600')
	for c in categories:
		worksheet.write(row, col, c, bold)
		if d[c][3]:
			worksheet.write(row+1, col, d[c][0], f)
			worksheet.write(row+2, col, d[c][1], f)
			worksheet.write(row+3, col, d[c][2], f)
		else:
			worksheet.write(row+1, col, d[c][0])
			worksheet.write(row+2, col, d[c][1])
			worksheet.write(row+3, col, d[c][2])
		col += 1
	workbook.close()
	
def avg_result_tables(filelist, categories, exp_name=''):
	'''
	Averages results tables
	'''
	micro_F1 = 0.0
	macro_F1 = 0.0
	n = len(filelist)
	d = dict()
	for c in categories:
		d[c] = [0.0, 0.0, 0.0, False]
	for f in filelist:
		name = 'results/'+f+'.xlsx'
		wb = load_workbook(filename = name)
		ws = wb.active
		micro_F1 += ws['A2'].value
		macro_F1 += ws['B2'].value
		col = 1
		letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
		for c in categories:
			d[c][0] += ws[letters[col]+'5'].value
			d[c][1] += ws[letters[col]+'6'].value
			d[c][2] += ws[letters[col]+'7'].value		
			col += 1
	for c in categories:
		d[c][0] = d[c][0]/n
		d[c][1] = d[c][1]/n
		d[c][2] = d[c][2]/n
	print_results(micro_F1/n, macro_F1/n, d, categories, exp_name='CV_'+exp_name)

